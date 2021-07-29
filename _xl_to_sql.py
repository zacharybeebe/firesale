from sqlite3 import connect
from openpyxl import load_workbook


CACHE = 'nfes_items.xlsx'
DB = 'NFES_ITEMS.db'

CACHES = ['AKK', 'BFK', 'CDK', 'GBK', 'LGK', 'LSK', 'NCK', 'NEK', 'NRK', 'NWK', 'PFK', 'RMK', 'SAK', 'SFK', 'WFK']

WS_DT = ['INT', 'TEXT', 'TEXT', 'TEXT', 'TEXT', 'TEXT', 'TEXT', 'REAL', 'TEXT', 'TEXT',
         'REAL', 'REAL', 'REAL', 'REAL', 'REAL', 'TEXT', 'TEXT', 'TEXT'] + (['BOOL'] * len(CACHES))



wb = load_workbook(CACHE, data_only=True)
ws = wb['ALL ITEMS']

master = []

for y, i in enumerate(ws.iter_rows()):
    if y != 0:
        if str(i[0].value)[0:2] == '00' and str(i[0].value)[0:2] != '08':
            lst = []
            for z, j in enumerate(i):
                if z == len(i) - 1:
                    lst.append(str(j.value))
                    caches = str(j.value).split(', ')
                    for c in CACHES:
                        if c in caches:
                            lst.append(1)
                        else:
                            lst.append(0)
                else:
                    if z in [4, 7, 8, 9, 10, 11]:
                        lst.append(float(j.value))
                    else:
                        if z in [0, 1, 2]:
                            lst.append(str(j.value).strip(' '))
                        else:
                            lst.append(str(j.value))
            if str(i[0].value[0]) == '0':
                if str(i[0].value[0:2]) != '08':
                    lst.insert(1, str(i[0].value).strip(' ')[2:])
            else:
                lst.insert(1, str(i[0].value).strip(' '))

            ty = str(i[2].value).split(' - ')[0].split(',')[0]
            if ty in ['JEAN', 'SHIRT']:
                lst.insert(2, 'NOMEX')
            else:
                lst.insert(2, ty)

            lst.insert(0, int(i[0].value))
            master.append(lst)



head = [f'"{i.value}"' for i in list(ws['1'])]
#head.pop(-1)
for c in CACHES:
    head.append(c)
head.insert(1, f'"NFES NO SHORT"')
head.insert(2, f'"TYPE"')
head.insert(0, f'"REF"')

print(head)


t_cols = f"""({', '.join([f'{a} {b}' for a, b in zip(head, WS_DT)])}, PRIMARY KEY ("REF"));"""



conn = connect(DB)
cur = conn.cursor()

t_sql = f"""CREATE TABLE items {t_cols}"""
t_idx = f"""CREATE Index IX_items_NFES_NO_SHORT ON items ("REF" ASC)"""

cur.execute(t_sql)
cur.execute(t_idx)
conn.commit()

for i in master:
    c_sql = f"""INSERT INTO items ({', '.join(head)}) VALUES ({', '.join(['?' for _ in i])})"""
    cur.execute(c_sql, i)

conn.commit()
conn.close()









