from sqlite3 import connect
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from pickle import (loads,
                    dumps)
from constants import (DB,
                       ORDER_INFO,
                       SQL_SEARCH_HEADER,
                       ITEM_COLS,
                       CART_COLS,
                       CACHE_INFO,
                       CACHE_NAMES,
                       USER_ARGS,
                       EXCEL_SQL_HEADER,
                       EXCEL_FILL)
from order_report import OrderReport


class Worker(object):
    def __init__(self, user, cache='all', nfes='all', cls='all', product='all', types='all', sort=0, prev_sort=1, reverse=0,
                 cart={}, cart_size=0, snum=0,report_info=deepcopy(ORDER_INFO), orders={}, order_num=1, lists={}):
        self.user = user
        self.cache = cache
        self.nfes = nfes
        self.cls = cls
        self.product = product
        self.types = types
        self.sort = sort
        self.prev_sort = prev_sort
        self.reverse = reverse
        self.cart = cart
        self.cart_size = cart_size
        self.snum = snum
        self.report_info = report_info
        self.orders = orders
        self.order_num = order_num
        self.lists = lists

        self.order_from_cache = None
        self.order_datetime = None

        self.check_date_for_snum()

    def __getitem__(self, item):
        return self.__dict__[item]

    def __setitem__(self, key, value):
        setattr(self, key, value)

    @classmethod
    def get_user(cls, user_db, user):
        conn, cur = Worker.connect_db(db=user_db)
        sql = f"""SELECT * FROM users WHERE user = ?"""
        cur.execute(sql, [user])
        info = cur.fetchone()
        conn.close()
        args = []
        for i in info:
            if isinstance(i, bytes):
                args.append(loads(i))
            else:
                args.append(i)
        return Worker(*args)

    @classmethod
    def delete_user(cls, user_db, user):
        conn, cur = Worker.connect_db(db=user_db)
        sql = f"""DELETE FROM users WHERE user = ?;"""
        cur.execute(sql, [user])
        conn.commit()
        conn.close()

    @staticmethod
    def connect_db(db=DB):
        connection = connect(db)
        cursor = connection.cursor()
        return connection, cursor

    @staticmethod
    def get_user_list(user_db):
        conn, cur = Worker.connect_db(db=user_db)
        sql = f"""SELECT user FROM users"""
        cur.execute(sql)
        x = [f'• {i[0]}' for i in cur.fetchall()]
        conn.close()
        return x

    @staticmethod
    def check_user(user_db, user):
        conn, cur = Worker.connect_db(db=user_db)
        sql = f"""SELECT user FROM users WHERE user = ?"""
        cur.execute(sql, [user])
        if cur.fetchone():
            conn.close()
            return True
        else:
            conn.close()
            return False

    @staticmethod
    def convert_binary(filepath):
        with open(filepath, 'rb') as file:
            binary = file.read()
        return binary

    @staticmethod
    def get_item_info(item):
        conn, cur = Worker.connect_db()
        sql = f"""SELECT {', '.join([f'"{key}"' for key in ITEM_COLS])} FROM items WHERE "REF" = ?"""
        cur.execute(sql, [item])
        x = cur.fetchone()
        conn.close()

        master = {}
        for i, key in enumerate(ITEM_COLS):
            if ITEM_COLS[key] == 'CUBIC FEET':
                master[ITEM_COLS[key]] = round(x[i], 4)
            else:
                master[ITEM_COLS[key]] = x[i]
        return master

    @staticmethod
    def format_date_time(dt):
        return f'{dt.month}/{dt.day}/{dt.year} {dt.hour}:{"0" * (2 - len(str(dt.minute))) + str(dt.minute)}'

    @staticmethod
    def format_date(dt):
        return f'{dt.month}/{dt.day}/{dt.year}'

    @staticmethod
    def format_date_report_filename(dt):
        return f'{dt.month}_{dt.day}_{dt.year}'

    @staticmethod
    def format_price(value):
        if isinstance(value, str):
            value = float(value)
        else:
            value = value
        val_list = [i for i in str(round(value, 2))]
        if '.' not in val_list:
            add_to = ['.', '0', '0']
            for i in add_to:
                val_list.append(i)
        else:
            if len(val_list[-(len(val_list) - val_list.index('.')):]) < 3:
                val_list.append('0')
        temp = [i for i in reversed(val_list)]
        added = 0
        for i in range(3, len(val_list)):
            if i != 3 and i % 3 == 0:
                temp.insert(i + added, ',')
                added += 1
        return '${}'.format(''.join([i for i in reversed(temp)]))

    def check_date_for_snum(self):
        if self.orders:
            if datetime.now().date() > max([self.orders[key]['datetime'].date() for key in self.orders]):
                self.snum = 0

    def check_cache_items(self, cache):
        items_in_cart = list(self.cart)
        conn, cur = Worker.connect_db()

        sql_in = f"""({', '.join([str(i) for i in items_in_cart])})"""
        sql = f"""SELECT {', '.join([f'"{i}"' for i in SQL_SEARCH_HEADER])} 
                  FROM items
                  WHERE "REF" IN {sql_in} AND "{cache}" = 0"""
        cur.execute(sql)
        items = cur.fetchall()
        conn.close()
        if items:
            return True, f'The following items are not supplied by the {CACHE_NAMES[cache]} Cache ({cache})', items
        else:
            return False, None, []

    def filter_items(self):
        items_in_cart = list(self.cart)
        conn, cur = Worker.connect_db()
        sql = f"""SELECT {', '.join([f'"{i}"' for i in SQL_SEARCH_HEADER])} FROM items"""

        if self.cache == 'all' and self.nfes == 'all' and self.cls == 'all' and self.product == 'all' and self.types == 'all':
            cur.execute(sql)
            items = [i for i in cur.fetchall() if int(i[0]) not in items_in_cart]
            return self.sort_items(items)
        else:
            sql += ' WHERE '
            add_and = False
            if self.cache != 'all':
                sql += f'"{self.cache}" = 1'
                add_and = True
            if self.nfes != 'all':
                if add_and:
                    sql += f' AND "REF" >= {int(self.nfes[0])} AND "REF" <= {int(self.nfes[1])}'
                else:
                    sql += f'"REF" >= {int(self.nfes[0])} AND "REF" <= {int(self.nfes[1])}'
                add_and = True
            if self.cls != 'all':
                if add_and:
                    sql += f' AND "CLASS" = "{self.cls}"'
                else:
                    sql += f'"CLASS" = "{self.cls}"'
                add_and = True
            if self.product != 'all':
                if add_and:
                    sql += f' AND "PROD LINE" = "{self.product}"'
                else:
                    sql += f'"PROD LINE" = "{self.product}"'
                add_and = True
            if self.types != 'all':
                if add_and:
                    sql += f' AND "TYPE" = "{self.types}"'
                else:
                    sql += f'"TYPE" = "{self.types}"'

            cur.execute(sql)
            items = [i for i in cur.fetchall() if int(i[0]) not in items_in_cart]
            conn.close()
            return self.sort_items(items)

    def sort_items(self, items):
        def get(idx, elem):
            if idx == 0:
                return int(elem[idx])
            else:
                return elem[idx]

        if self.sort == self.prev_sort:
            rev = int(not bool(int(self.reverse)))
        else:
            rev = 0
        lst = sorted(items, key=lambda x: get(int(self.sort), x), reverse=bool(rev))
        self.reverse = rev
        return lst

    def add_cart(self, item, qty=1):
        conn, cur = Worker.connect_db()
        sql = f"""SELECT {', '.join([f'"{key}"' for key in CART_COLS])} FROM items WHERE "REF" = ?"""
        cur.execute(sql, [item])
        x = cur.fetchone()
        conn.close()

        master = {}
        for i, key in enumerate(CART_COLS):
            if CART_COLS[key] == 'PRICE PER ITEM':
                master[CART_COLS[key]] = x[i]
                master['PRICE TOTAL'] = x[i] * qty
            else:
                master[CART_COLS[key]] = x[i]
        master['QUANTITY'] = qty
        self.cart.update({item: master})

    def add_list(self, list_, item, qty=1):
        conn, cur = Worker.connect_db()
        sql = f"""SELECT {', '.join([f'"{key}"' for key in CART_COLS])} FROM items WHERE "REF" = ?"""
        cur.execute(sql, [item])
        x = cur.fetchone()
        conn.close()

        master = {}
        for i, key in enumerate(CART_COLS):
            if CART_COLS[key] == 'PRICE PER ITEM':
                master[CART_COLS[key]] = x[i]
                master['PRICE TOTAL'] = x[i] * qty
            else:
                master[CART_COLS[key]] = x[i]
        master['QUANTITY'] = qty
        self.lists[list_]['cart'].update({item: master})
        self.lists[list_]['total_price'] = self.list_total_price(list_)
        self.lists[list_]['total_qty'] = self.list_total_qty(list_)

    def update_cart_size(self):
        self.cart_size = len(self.cart)

    def cart_total_price(self):
        return sum([self.cart[key]['PRICE TOTAL'] for key in self.cart])

    def cart_total_qty(self):
        return sum([self.cart[key]['QUANTITY'] for key in self.cart])

    def add_new_list(self, list_):
        self.lists[list_] = {'datetime': datetime.now(),
                             'total_price': 0,
                             'total_qty': 0,
                             'cart': {}}

    def list_total_price(self, list_):
        return sum([self.lists[list_]['cart'][key]['PRICE TOTAL'] for key in self.lists[list_]['cart']])

    def list_total_qty(self, list_):
        return sum([self.lists[list_]['cart'][key]['QUANTITY'] for key in self.lists[list_]['cart']])

    def list_again(self, list_):
        for key in self.lists[list_]['cart']:
            i_num = int(key)
            self.add_cart(i_num, qty=self.lists[list_]['cart'][i_num]['QUANTITY'])
        self.update_cart_size()

    def reorder_list(self, list_):
        sort = sorted(self.lists[list_]['cart'])
        master = {}
        for i_num in sort:
            master[i_num] = {}
            master[i_num].update(self.lists[list_]['cart'][i_num])
        return master

    def reorder_snums(self):
        sort = sorted(list(self.cart))
        master = {}
        for i, i_num in enumerate(sort):
            master[i_num] = {'S #': f"""S-{'0' * (4 - len(str(self.snum + i + 1)))}{self.snum + i + 1}"""}
            master[i_num].update(self.cart[i_num])
        return master

    def publish_snums(self):
        sort = sorted(list(self.cart))
        master = {}
        for i_num in sort:
            master[i_num] = {'S #': f"""S-{'0' * (4 - len(str(self.snum + 1)))}{self.snum + 1}"""}
            master[i_num].update(self.cart[i_num])
            self.snum += 1
        return master

    def order_again(self, order):
        for key in self.orders[order]['cart']:
            i_num = int(key)
            self.add_cart(i_num, qty=self.orders[order]['cart'][i_num]['QUANTITY'])
        self.update_cart_size()

    def file_order(self):
        key = self.report_info['1Incident Order Number']
        self.orders[key] = {}
        self.orders[key]['datetime'] = self.order_datetime
        self.orders[key]['total_price'] = self.cart_total_price()
        self.orders[key]['total_qty'] = self.cart_total_qty()
        self.orders[key]['cart'] = self.publish_snums()
        self.orders[key]['order_info'] = deepcopy(self.report_info)
        self.orders[key]['report_name'] = f'order_{key}_{self.format_date_report_filename(self.order_datetime)}.pdf'
        self.orders[key]['excel_name'] = f'order_{key}_{self.format_date_report_filename(self.order_datetime)}.xlsx'
        self.create_order(key)
        self.create_excel(key)
        self.report_info = deepcopy(ORDER_INFO)

    def fill_cache_info(self):
        for info in CACHE_INFO[self.order_from_cache]:
            self.report_info[info] = str(CACHE_INFO[self.order_from_cache][info])
        self.report_info['1Incident Order Number'] = f"""{'0' * (3 - len(str(self.order_num)))}{self.order_num}"""
        self.order_datetime = datetime.now()
        self.report_info['9Date/Time Ordered'] = self.format_date_time(self.order_datetime)

    def create_order(self, order):
        report = OrderReport()
        report.alias_nb_pages()
        report.add_page()
        report.compile_report(self.orders[order]['cart'], self.orders[order]['order_info'])
        self.orders[order]['report'] = report.output(dest='S').encode('latin-1')

    def create_excel(self, order):
        conn, cur = self.connect_db()
        sql_cols = f"""{', '.join([i[0] for i in EXCEL_SQL_HEADER])}"""
        sql_in = f"""({', '.join([str(key) for key in self.orders[order]['cart']])})"""
        sql = f"""SELECT {sql_cols} FROM items WHERE "REF" IN {sql_in}"""
        cur.execute(sql)
        items = cur.fetchall()
        conn.close()

        temp = {}
        for i in items:
            for j, head_col in zip(i, EXCEL_SQL_HEADER):
                if head_col[1] not in temp:
                    temp[head_col[1]] = []
                temp[head_col[1]].append(j)
        for key in self.orders[order]['cart']:
            for head_col in EXCEL_FILL:
                if head_col[1] not in temp:
                    temp[head_col[1]] = []
                temp[head_col[1]].append(self.orders[order]['cart'][key][head_col[0]])

        master = {key: temp[key] for key in sorted(temp)}
        heads = [[i[0].strip('""'), i[1]] for i in sorted(EXCEL_SQL_HEADER + EXCEL_FILL, key=lambda x: x[1])]
        wb = Workbook()
        ws = wb.active
        ws.title = self.orders[order]['excel_name'][:-5]
        for head in heads:
            ws.cell(1, head[1], head[0])
            for i, item in enumerate(master[head[1]]):
                ws.cell(i+2, head[1], item)
        self.orders[order]['excel'] = save_virtual_workbook(wb)

    def download_order_pdf(self, order):
        return BytesIO(self.orders[order]['report']), self.orders[order]['report_name']

    def download_order_excel(self, order):
        return BytesIO(self.orders[order]['excel']), self.orders[order]['excel_name']

    def download_all_by_item_excel(self):
        all_items = []
        for order in self.orders:
            for item in self.orders[order]['cart']:
                all_items.append(item)
        all_items = sorted({i for i in all_items})

        conn, cur = self.connect_db()
        sql_cols = f"""{', '.join([i[0] for i in EXCEL_SQL_HEADER])}"""
        sql_in = f"""({', '.join([str(i) for i in all_items])})"""
        sql = f"""SELECT {sql_cols} FROM items WHERE "REF" IN {sql_in}"""
        cur.execute(sql)
        items = {int(i[0]): [[j, z[1]] for j, z in zip(i, EXCEL_SQL_HEADER)] for i in cur.fetchall()}
        conn.close()

        for key in EXCEL_FILL[1:]:
            for item in items:
                add = [0, key[1]]
                for order in self.orders:
                    if item in self.orders[order]['cart']:
                        add[0] += self.orders[order]['cart'][item][key[0]]
                items[item].append(add)
        for item in items:
            items[item] = sorted(items[item], key=lambda x: x[1])

        heads = [[i[0].strip('""'), i[1]] for i in sorted(EXCEL_SQL_HEADER + EXCEL_FILL[1:], key=lambda x: x[1])]

        file_name = 'all_orders_by_item.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = file_name[:-5]
        for head in heads:
            ws.cell(1, head[1] - 1, head[0])
        for i, item in enumerate(items):
            for dat in items[item]:
                ws.cell(i + 2, dat[1] - 1, dat[0])

        return BytesIO(save_virtual_workbook(wb)), file_name

    def download_all_by_orders_excel(self):
        conn, cur = self.connect_db()
        wb = Workbook()
        for index, order in enumerate(self.orders):
            sql_cols = f"""{', '.join([i[0] for i in EXCEL_SQL_HEADER])}"""
            sql_in = f"""({', '.join([str(key) for key in self.orders[order]['cart']])})"""
            sql = f"""SELECT {sql_cols} FROM items WHERE "REF" IN {sql_in}"""
            cur.execute(sql)
            items = cur.fetchall()

            temp = {}
            for i in items:
                for j, head_col in zip(i, EXCEL_SQL_HEADER):
                    if head_col[1] not in temp:
                        temp[head_col[1]] = []
                    temp[head_col[1]].append(j)
            for key in self.orders[order]['cart']:
                for head_col in EXCEL_FILL:
                    if head_col[1] not in temp:
                        temp[head_col[1]] = []
                    temp[head_col[1]].append(self.orders[order]['cart'][key][head_col[0]])

            master = {key: temp[key] for key in sorted(temp)}
            heads = [[i[0].strip('""'), i[1]] for i in sorted(EXCEL_SQL_HEADER + EXCEL_FILL, key=lambda x: x[1])]

            sheet_name = self.orders[order]['excel_name'][:-5]
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)
            for head in heads:
                ws.cell(1, head[1], head[0])
                for i, item in enumerate(master[head[1]]):
                    ws.cell(i + 2, head[1], item)
        conn.close()
        return BytesIO(save_virtual_workbook(wb)), 'all_orders_by_order.xlsx'

    def reset_variables(self):
        self.order_num += 1
        self.cart = {}
        self.cart_size = 0

    def insert_self(self, user_db):
        conn, cur = self.connect_db(db=user_db)
        d = self.__dict__
        vals = []
        for key in d:
            if key not in ['order_from_cache', 'order_datetime']:
                if USER_ARGS[key] == 'BLOB':
                    vals.append(dumps(d[key]))
                else:
                    vals.append(d[key])
        sql = f"""INSERT into users ({', '.join(list(USER_ARGS))}) VALUES ({', '.join(['?' for _ in vals])})"""
        cur.execute(sql, vals)
        conn.commit()
        conn.close()

    def update_self(self, user_db):
        conn, cur = self.connect_db(db=user_db)
        d = self.__dict__
        vals = []
        for key in d:
            if key not in ['order_from_cache', 'order_datetime']:
                if USER_ARGS[key] == 'BLOB':
                    vals.append(dumps(d[key]))
                else:
                    vals.append(d[key])
        sql = f"""UPDATE users 
                  SET ({', '.join(list(USER_ARGS))}) = ({', '.join(['?' for _ in vals])}) 
                  WHERE user = ?;"""
        vals.append(self.user)
        cur.execute(sql, vals)
        conn.commit()
        conn.close()






if __name__ == '__main__':
    pass
    # w = Worker('g')
    # for i in ['fire', 'test']:
    #     Worker.delete_user(i)
    # sql_cols = f"""({', '.join([f'{key} {USER_ARGS[key]}' for key in USER_ARGS])})"""
    # print(sql_cols)
    #
    # conn, cur = w.connect_db(db=UDB)
    #
    # sql = f"""CREATE TABLE users {sql_cols};"""
    # cur.execute(sql)
    # conn.commit()
    # conn.close()
